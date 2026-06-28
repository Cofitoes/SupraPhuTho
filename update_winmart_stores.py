import pandas as pd
import json
import time
import requests
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"G:\My Drive\Training AI\Supra Phú Thọ\DSCuaHangFinal.xlsx"
js_path = r"G:\My Drive\Training AI\Supra Phú Thọ\store_data.js"

def check_and_update_excel():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Read header row (row 1)
        headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        headers_lower = [str(h).lower().strip() if h else "" for h in headers]
        
        # Find column indices (1-indexed for openpyxl)
        def find_col(options):
            for opt in options:
                for idx, h in enumerate(headers_lower):
                    if h == opt.lower().strip():
                        return idx + 1
            return -1
            
        id_idx = find_col(["Site Store", "Mã cửa hàng (Mã CH)", "Mã cửa hàng", "ID", "Store_ID"])
        name_idx = find_col(["Tên cửa hàng", "Name", "Store_Name"])
        addr_idx = find_col(["địa chỉ", "Địa chỉ", "Address"])
        prov_idx = find_col(["Tỉnh giao", "Thành Phố/Tỉnh", "Tỉnh", "Province", "Tên tỉnh chuẩn"])
        dist_idx = find_col(["District", "Quận/Huyện", "Huyện/Xã", "Huyện"])
        type_idx = find_col(["Trip_Type", "Phân loại", "Trip Type"])
        lat_idx = find_col(["Lat", "Vi Do", "Vĩ độ"])
        lng_idx = find_col(["Long", "Lng", "Kinh Do", "Kinh độ"])
        
        # Ensure we have required columns
        if id_idx == -1 or name_idx == -1 or addr_idx == -1:
            print("Required columns missing in Excel headers")
            wb.close()
            return
            
        # Check if store already exists by ID or name
        exists = False
        target_name = "WM+ PTO 33 Thống Nhất, Phùng Nguyên"
        target_id = "2AKG"
        
        for r in range(2, sheet.max_row + 1):
            name_val = sheet.cell(row=r, column=name_idx).value
            id_val = sheet.cell(row=r, column=id_idx).value
            
            # Check if name matches or ID matches target_id and name is similar
            if (name_val and str(name_val).strip() == target_name) or (id_val and str(id_val).strip() == target_id and name_val and "Thống Nhất" in str(name_val)):
                exists = True
                # Update store name and coords
                sheet.cell(row=r, column=name_idx, value=target_name)
                if lat_idx != -1: sheet.cell(row=r, column=lat_idx, value=21.2819416)
                if lng_idx != -1: sheet.cell(row=r, column=lng_idx, value=105.3053024)
                print(f"Updated store coordinates at row {r}")
                break
                
        if not exists:
            # Append store
            r = sheet.max_row + 1
            print(f"Appending store to row {r}")
            sheet.cell(row=r, column=id_idx, value=target_id)
            sheet.cell(row=r, column=name_idx, value=target_name)
            sheet.cell(row=r, column=addr_idx, value="Số 33 Thống Nhất, Phùng Nguyên, H. Lâm Thao, T. Phú Thọ")
            if prov_idx != -1: sheet.cell(row=r, column=prov_idx, value="T. Phú Thọ")
            if dist_idx != -1: sheet.cell(row=r, column=dist_idx, value="H. Lâm Thao")
            if type_idx != -1: sheet.cell(row=r, column=type_idx, value="Giao Thang")
            if lat_idx != -1: sheet.cell(row=r, column=lat_idx, value=21.2819416)
            if lng_idx != -1: sheet.cell(row=r, column=lng_idx, value=105.3053024)
            
        wb.save(file_path)
        wb.close()
        print("Excel check and update completed successfully.")
    except Exception as e:
        print(f"Error checking Excel: {e}")

check_and_update_excel()

print("Loading DSCuaHangFinal.xlsx...")
df = pd.read_excel(file_path, engine="openpyxl")

# Find index of columns
def get_col(df, options):
    for opt in options:
        for c in df.columns:
            if str(c).lower().strip() == opt.lower():
                return c
    return None

id_col = get_col(df, ["Site Store", "Mã cửa hàng (Mã CH)", "Mã cửa hàng", "ID", "Store_ID"])
name_col = get_col(df, ["Tên cửa hàng", "Name", "Store_Name"])
addr_col = get_col(df, ["địa chỉ", "Địa chỉ", "Address"])
prov_col = get_col(df, ["Tỉnh giao", "Thành Phố/Tỉnh", "Tỉnh", "Province", "Tên tỉnh chuẩn"])
dist_col = get_col(df, ["District", "Quận/Huyện", "Huyện/Xã", "Huyện"])
type_col = get_col(df, ["Trip_Type", "Phân loại", "Trip Type"])
lat_col = get_col(df, ["Lat", "Vi Do", "Vĩ độ"])
lng_col = get_col(df, ["Long", "Lng", "Kinh Do", "Kinh độ"])
latlong_col = get_col(df, ["Vị trí Lat/Long", "LatLong", "Lat/Long"])

if not all([id_col, name_col, addr_col]):
    print("Error: Could not find required columns in Excel.")
    sys.exit(1)

store_list = []
updated = 0

print(f"Columns matched: ID={id_col}, Name={name_col}, Addr={addr_col}, Prov={prov_col}, Dist={dist_col}, Type={type_col}, Lat={lat_col}, Lng={lng_col}, LatLong={latlong_col}")

for idx, row in df.iterrows():
    store_id = str(row[id_col]).strip()
    if store_id == 'nan' or not store_id:
        continue
        
    name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
    address = str(row[addr_col]).strip() if pd.notna(row[addr_col]) else ""
    prov = str(row[prov_col]).strip() if prov_col and pd.notna(row[prov_col]) else ""
    district = str(row[dist_col]).strip() if dist_col and pd.notna(row[dist_col]) else ""
    trip_type = str(row[type_col]).strip() if type_col and pd.notna(row[type_col]) else ""
    
    lat = 0.0
    lng = 0.0
    
    # Try separate Lat/Long columns first
    if lat_col and lng_col:
        lat = float(row[lat_col]) if pd.notna(row[lat_col]) else 0.0
        lng = float(row[lng_col]) if pd.notna(row[lng_col]) else 0.0
    
    # If no separate columns or both are 0, try combined Lat/Long column
    if (lat == 0.0 or lng == 0.0) and latlong_col and pd.notna(row.get(latlong_col)):
        latlong_str = str(row[latlong_col]).strip()
        if ',' in latlong_str:
            parts = latlong_str.split(',')
            try:
                lat = float(parts[0].strip())
                lng = float(parts[1].strip())
            except (ValueError, IndexError):
                pass
    
    # Check if we need geocoding
    if lat == 0.0 or lng == 0.0 or pd.isna(lat) or pd.isna(lng):
        print(f"Geocoding {store_id}: {address}")
        # Simplistic geocoding logic
        q = f"{address}"
        try:
            headers = {"User-Agent": "LogisticsHub-Winmart/1.0"}
            r = requests.get(f"https://nominatim.openstreetmap.org/search?format=json&q={q}&countrycodes=vn", headers=headers, timeout=10)
            data = r.json()
            if data and len(data) > 0:
                lat = float(data[0]['lat'])
                lng = float(data[0]['lon'])
                print(f"  -> Found: {lat}, {lng}")
                updated += 1
            else:
                print(f"  -> Not found")
        except Exception as e:
            print(f"  -> Error: {e}")
        time.sleep(1.2) # Nominatim rate limit
        
    coords = None
    if lat and lng and lat != 0.0 and lng != 0.0:
        coords = {"lat": lat, "lng": lng}
        
    # Apply manual overrides for known problematic stores
    MANUAL_OVERRIDES = {
        "WM+ PTO Khu Suông 2, Phú Khê": {
            "coords": {"lat": 21.387993, "lng": 105.087082},
            "id": "2CGH"
        },
        "WM+ PTO Khu 8, Tiêu Sơn, Chân Mộng": {
            "coords": {"lat": 21.56417, "lng": 105.17778},
            "id": "2CFI"
        },
        "WM+ PTO Khu 3, Liên Minh": {
            "coords": {"lat": 21.3753828, "lng": 105.1860164},
            "id": "2CCB"
        },
        "WM+ PTO KĐT Âu Cơ, Âu Cơ": {
            "coords": {"lat": 21.3989531, "lng": 105.2072326},
            "id": "2CFK"
        },
        "WM VC+ PTO Phú Thọ": {
            "coords": {"lat": 21.4184405, "lng": 105.2118986},
            "id": "1649"
        },
        "WM+ PTO Khu 21, Vạn Xuân": {
            "coords": {"lat": 21.31667, "lng": 105.26444},
            "id": "2ANC"
        },
        "WM+ PTO Khu 1, Thanh Thủy": {
            "coords": {"lat": 21.134, "lng": 105.280},
            "id": "2ALI"
        },
        "WM+ PTO Khu Phố, TT Thanh Thủy": {
            "coords": {"lat": 21.1705, "lng": 105.279},
            "id": "2APX"
        },
        "WM+ PTO Khu 4, Đoan Hạ": {
            "coords": {"lat": 21.13444, "lng": 105.27972},
            "id": "2AKU"
        },
        "WM+ PTO 33 Thống Nhất, Phùng Nguyên": {
            "coords": {"lat": 21.2819416, "lng": 105.3053024},
            "id": "2AKG"
        },
        "WM+ PTO 33 Thống Nhất Phùng Nguyên": {
            "coords": {"lat": 21.2819416, "lng": 105.3053024},
            "id": "2AKG"
        },
        # Add future overrides here
    }
    
    if name in MANUAL_OVERRIDES:
        coords = MANUAL_OVERRIDES[name]["coords"]
        
    store_list.append({
        "id": store_id,
        "name": name,
        "address": address,
        "province": prov,
        "district": district,
        "trip_type": trip_type,
        "isGXT": trip_type == "GXT",
        "coords": coords
    })

# Add missing manual overrides that weren't in the Excel
found_names = {s['name'] for s in store_list}
for o_name, o_data in MANUAL_OVERRIDES.items():
    if o_name not in found_names:
        is_gxt = "Phú Thọ" in o_name or "Việt Trì" in o_name or o_data.get("id") in ["1649", "1564"]
        store_list.append({
            "id": o_data.get("id", o_name),
            "name": o_name,
            "address": o_name,
            "province": "Phú Thọ",
            "district": "TX. Phú Thọ" if "Phú Thọ" in o_name else "TP. Việt Trì",
            "trip_type": "GXT" if is_gxt else "Giao Thang",
            "isGXT": is_gxt,
            "coords": o_data["coords"]
        })
        print(f"Injected missing store from overrides: {o_name}")

print(f"Total stores extracted: {len(store_list)}")
print(f"Newly geocoded: {updated}")

with open(js_path, "w", encoding="utf-8") as f:
    f.write(f"const STORE_LIST_DATA = {json.dumps(store_list, ensure_ascii=False, indent=2)};\n")
print("store_data.js generated successfully.")
