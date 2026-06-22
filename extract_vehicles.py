import openpyxl
import os
import glob
import re
import json
import sys

# Đảm bảo mã hóa UTF-8 cho dòng ra
sys.stdout.reconfigure(encoding='utf-8')

output_js_path = r"g:\My Drive\Training AI\Supra Phú Thọ\vehicle_data.js"
booking_dir = r"g:\My Drive\Training AI\Supra Phú Thọ\Data_Booking"

def main():
    extracted_data = {}
    
    # Tìm kiếm tất cả các tệp Planning*.xlsx
    pattern = os.path.join(booking_dir, "Planning*.xlsx")
    files = glob.glob(pattern)
    print(f"Tìm thấy {len(files)} tệp planning để phân tích.")
    
    for file_path in files:
        filename = os.path.basename(file_path)
        # Trích xuất ngày booking từ tên tệp linh hoạt
        date_str = None
        # Định dạng chuẩn: Planning_YYYY-MM-DD.xlsx
        match_std = re.search(r'Planning_(\d{4}-\d{2}-\d{2})\.xlsx', filename)
        if match_std:
            date_str = match_std.group(1)
        else:
            # Định dạng khác: Planning tpt CC-GHN (ord) CJ MQ 9.06.2026.xlsx, etc.
            match_alt1 = re.search(r'(\d{1,2})[._-](\d{1,2})[._-](\d{4})', filename)
            if match_alt1:
                day = int(match_alt1.group(1))
                month = int(match_alt1.group(2))
                year = int(match_alt1.group(3))
                date_str = f"{year:04d}-{month:02d}-{day:02d}"
            else:
                match_alt2 = re.search(r'(\d{4})[._-](\d{1,2})[._-](\d{1,2})', filename)
                if match_alt2:
                    year = int(match_alt2.group(1))
                    month = int(match_alt2.group(2))
                    day = int(match_alt2.group(3))
                    date_str = f"{year:04d}-{month:02d}-{day:02d}"
                    
        if not date_str:
            print(f"Bỏ qua tệp {filename} vì không phân tích được ngày.")
            continue
        print(f"Đang phân tích tệp {filename} cho ngày {date_str}...")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb["Form"] if "Form" in wb.sheetnames else wb.active
            
            # Hủy gộp ô để đọc chính xác
            if sheet.merged_cells.ranges:
                for r in list(sheet.merged_cells.ranges):
                    sheet.unmerge_cells(str(r))
            
            headers = [cell.value for cell in sheet[2]]
            if not headers or "Ship code" not in headers:
                # Try sheet[1] in case headers are on row 1
                headers = [cell.value for cell in sheet[1]]
                if not headers or "Ship code" not in headers:
                    print(f"Skipping {filename}: 'Ship code' header not found.")
                    wb.close()
                    continue
            
            if "Trip" not in headers or "Truck Number" not in headers:
                print(f"Skipping {filename}: 'Trip' or 'Truck Number' header not found.")
                wb.close()
                continue
                
            trip_col = headers.index("Trip")
            plate_col = headers.index("Truck Number")
            driver_col = headers.index("Driver Name") if "Driver Name" in headers else 12
            phone_col = 13 # Dự phòng vị trí cột N (index 13) cho SĐT tài xế
            
            vehicles = {}
            for r_idx in range(3, sheet.max_row + 1):
                trip_val = sheet.cell(row=r_idx, column=trip_col+1).value
                plate_val = sheet.cell(row=r_idx, column=plate_col+1).value
                driver_val = sheet.cell(row=r_idx, column=driver_col+1).value
                phone_val = sheet.cell(row=r_idx, column=phone_col+1).value
                
                if trip_val is not None:
                    try:
                        trip_num = str(int(float(trip_val)))
                    except:
                        trip_num = str(trip_val).strip()
                        
                    plate = str(plate_val).strip() if plate_val else ""
                    driver = str(driver_val).strip() if driver_val else ""
                    
                    phone = ""
                    if phone_val is not None:
                        try:
                            # Chuẩn hóa định dạng số điện thoại
                            phone_str = str(int(float(phone_val)))
                            if not phone_str.startswith("0") and len(phone_str) >= 9:
                                phone = "0" + phone_str
                            else:
                                phone = phone_str
                        except:
                            phone = str(phone_val).strip()
                            
                    if trip_num and plate:
                        if trip_num not in vehicles:
                            vehicles[trip_num] = {
                                "plate": plate,
                                "driver": driver,
                                "phone": phone
                            }
            
            if vehicles:
                extracted_data[date_str] = vehicles
                print(f" -> Đã trích xuất thành công {len(vehicles)} xe cho ngày {date_str}.")
            wb.close()
        except Exception as e:
            print(f"Lỗi khi đọc tệp {filename}: {e}")
            
    # Ghi vào file vehicle_data.js trong thư mục workspace
    js_content = f"const VEHICLE_DATA = {json.dumps(extracted_data, ensure_ascii=False, indent=2)};\n"
    with open(output_js_path, "w", encoding="utf-8") as f:
        f.write(js_content)
    print(f"Đã tạo thành công {output_js_path} với {len(extracted_data)} ngày dữ liệu.")

if __name__ == "__main__":
    main()
