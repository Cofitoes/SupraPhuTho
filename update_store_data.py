import json
import os
import sys
import openpyxl
import urllib.request
import re
import datetime
import unicodedata
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

# Paths
folder_path = os.path.dirname(os.path.abspath(__file__))
LOCAL_EXCEL_PATH = os.path.join(folder_path, "DS_Cua_Hang.xlsm")
if not os.path.exists(LOCAL_EXCEL_PATH):
    alt_path1 = os.path.join(folder_path, "DS_Cua_Hang_Updated.xlsx")
    alt_path2 = os.path.join(folder_path, "DS_Cua_Hang.xlsx")
    if os.path.exists(alt_path1):
        LOCAL_EXCEL_PATH = alt_path1
    elif os.path.exists(alt_path2):
        LOCAL_EXCEL_PATH = alt_path2

gsheet_xlsx_path = os.path.join(folder_path, "DS_Cua_Hang_GSheet.xlsx")
updated_xlsx_path = os.path.join(folder_path, "DS_Cua_Hang_Updated.xlsx")
concung_json_path = os.path.join(folder_path, "concung_stores.json")
store_data_js_path = os.path.join(folder_path, "store_data.js")

# Define Northern provinces for region classification
NORTH_PROVINCES = {
    "HÃ  Ná»™i", "Háº£i PhÃ²ng", "Quáº£ng Ninh", "Báº¯c Giang", "Báº¯c Ninh", "HÃ  Nam", 
    "HÃ  TÄ©nh", "HÆ°ng YÃªn", "Háº£i DÆ°Æ¡ng", "Lai ChÃ¢u", "LÃ o Cai", "Láº¡ng SÆ¡n", 
    "Nam Äá»‹nh", "Nghá»‡ An", "Ninh BÃ¬nh", "Quáº£ng Trá»‹", "Thanh HÃ³a", "Thanh HoÃ¡", 
    "ThÃ¡i BÃ¬nh", "ThÃ¡i NguyÃªn", "VÄ©nh PhÃºc", "PhÃº Thá»", "TuyÃªn Quang", "YÃªn BÃ¡i", 
    "Thá»«a ThiÃªn Huáº¿", "HÃ  Giang", "Cao Báº±ng", "Báº¯c Káº¡n", "SÆ¡n La", "HÃ²a BÃ¬nh", "Äiá»‡n BiÃªn"
}

PREFIX_MAP = {
    'AGI': 'An Giang', 'BDI': 'BÃ¬nh Äá»‹nh', 'BDU': 'BÃ¬nh DÆ°Æ¡ng', 'BGI': 'Báº¯c Giang', 'BLI': 'Báº¡c LiÃªu',
    'BNI': 'Báº¯c Ninh', 'BPH': 'BÃ¬nh PhÆ°á»›c', 'BTH': 'BÃ¬nh Thuáº­n', 'BTR': 'Báº¿n Tre', 'BVT': 'BÃ  Rá»‹a - VÅ©ng TÃ u',
    'CMA': 'CÃ  Mau', 'CTO': 'Cáº§n ThÆ¡', 'DLK': 'Äáº¯k Láº¯k', 'DNA': 'ÄÃ  Náºµng', 'DAN': 'ÄÃ  Náºµng', 'DNG': 'ÄÃ  Náºµng',
    'DNK': 'Äáº¯k NÃ´ng', 'DON': 'Äá»“ng Nai', 'GLA': 'Gia Lai', 'HAG': 'Háº­u Giang', 'HAP': 'Háº£i PhÃ²ng',
    'HCM': 'Há»“ ChÃ­ Minh', 'HDU': 'Háº£i DÆ°Æ¡ng', 'HGI': 'Háº­u Giang', 'HNA': 'HÃ  Nam', 'HNI': 'HÃ  Ná»™i',
    'HOB': 'HÃ²a BÃ¬nh', 'HTI': 'HÃ  TÄ©nh', 'HUE': 'Thá»«a ThiÃªn Huáº¿', 'HYE': 'HÆ°ng YÃªn', 'HYN': 'HÆ°ng YÃªn',
    'KGI': 'KiÃªn Giang', 'KHO': 'KhÃ¡nh HÃ²a', 'KTU': 'Kon Tum', 'LAN': 'Long An', 'LCA': 'LÃ o Cai',
    'LCH': 'Lai ChÃ¢u', 'LDO': 'LÃ¢m Äá»“ng', 'LSO': 'Láº¡ng SÆ¡n', 'NAN': 'Nghá»‡ An', 'NBI': 'Ninh BÃ¬nh',
    'NDI': 'Nam Äá»‹nh', 'NTH': 'Ninh Thuáº­n', 'PTH': 'PhÃº Thá»', 'PYN': 'PhÃº YÃªn', 'QBI': 'Quáº£ng BÃ¬nh',
    'QNA': 'Quáº£ng Nam', 'QNG': 'Quáº£ng NgÃ£i', 'QNI': 'Quáº£ng Ninh', 'QTR': 'Quáº£ng Trá»‹', 'SLA': 'SÆ¡n La',
    'STR': 'SÃ³c TrÄƒng', 'TBI': 'ThÃ¡i BÃ¬nh', 'TGI': 'Tiá»n Giang', 'THO': 'Thanh HÃ³a', 'TNG': 'ThÃ¡i NguyÃªn',
    'TNI': 'TÃ¢y Ninh', 'TQU': 'TuyÃªn Quang', 'TVI': 'TrÃ  Vinh', 'VLO': 'VÄ©nh Long', 'VPH': 'VÄ©nh PhÃºc',
    'YBA': 'YÃªn BÃ¡i'
}

def clean_province_name(prov_name, store_name=None):
    # 1. Resolve via store name prefix if possible
    if store_name:
        store_name_clean = store_name.strip()
        prefix_m = re.match(r"^([A-Z]{3,4})", store_name_clean.upper())
        if prefix_m:
            prefix = prefix_m.group(1)
            pref3 = prefix[:3]
            if pref3 in PREFIX_MAP:
                prov_name_mapped = PREFIX_MAP[pref3]
                if prov_name_mapped in ["Há»“ ChÃ­ Minh", "HÃ  Ná»™i", "ÄÃ  Náºµng", "Háº£i PhÃ²ng", "Cáº§n ThÆ¡"]:
                    return f"ThÃ nh Phá»‘ {prov_name_mapped}"
                else:
                    return f"Tá»‰nh {prov_name_mapped}"

    if not prov_name:
        return ""
    
    # 2. Normalize unicode to NFC
    prov = unicodedata.normalize('NFC', str(prov_name)).strip()
    
    # 3. Remove parentheses content like (coming soon ...)
    prov = re.sub(r'\(.*?\)', '', prov).strip()
    
    # 4. Remove trailing periods or spaces
    prov = prov.strip(". ")
    
    # 5. Extract province name if contains "tá»‰nh"
    if "tá»‰nh" in prov.lower():
        parts = re.split(r'(?i)\btá»‰nh\b', prov)
        if len(parts) > 1:
            prov = parts[-1].strip()
            
    # 6. Remove standard administrative prefixes from the raw string
    prov = re.sub(r'(?i)^(ThÃ nh Phá»‘|ThÃ nh phá»‘|thÃ nh phá»‘|Tá»‰nh|tp\.|tp|TP\.|TP)\s+', '', prov).strip()
    
    # 7. Specific manual corrections
    prov_lower = prov.lower()
    if prov_lower == "huáº¿":
        prov = "Thá»«a ThiÃªn Huáº¿"
    elif prov_lower == "bÃ n tháº¡ch":
        prov = "Quáº£ng Nam"
    elif prov_lower == "viá»‡t nam" or prov_lower == "vietnam":
        prov = "Há»“ ChÃ­ Minh"
        
    # 8. Standardize with "ThÃ nh Phá»‘ " or "Tá»‰nh " prefix
    if prov in ["Há»“ ChÃ­ Minh", "HÃ  Ná»™i", "ÄÃ  Náºµng", "Háº£i PhÃ²ng", "Cáº§n ThÆ¡"]:
        return f"ThÃ nh Phá»‘ {prov}"
    else:
        return f"Tá»‰nh {prov}"


def is_northern_province(prov_name):
    if not prov_name:
        return False
    clean_prov = prov_name.replace("ThÃ nh Phá»‘", "").replace("ThÃ nh phá»‘", "").replace("thÃ nh phá»‘", "")
    clean_prov = clean_prov.replace("Tá»‰nh", "").replace("tá»‰nh", "").replace("TP.", "").replace("TP", "")
    clean_prov = clean_prov.strip(". ")
    return any(p in clean_prov for p in NORTH_PROVINCES)

def excel_datetime_to_float(dt_val):
    if isinstance(dt_val, (datetime.datetime, datetime.date)):
        if dt_val.month in [8, 9, 10, 11, 12]:
            if dt_val.day == 1 and dt_val.year >= 2000:
                year_str = str(dt_val.year)
                return dt_val.month + float(dt_val.year) / (10 ** len(year_str))
            elif dt_val.day != 1 and 2000 <= dt_val.year <= 2030:
                day_str = str(dt_val.day)
                return dt_val.month + float(dt_val.day) / (10 ** len(day_str))
                
        # Standard conversion using serial numbers
        if isinstance(dt_val, datetime.datetime):
            if dt_val < datetime.datetime(1900, 3, 1):
                epoch = datetime.datetime(1899, 12, 31)
            else:
                epoch = datetime.datetime(1899, 12, 30)
            delta = dt_val - epoch
            return delta.days + (delta.seconds + delta.microseconds / 1e6) / 86400.0
        else: # datetime.date
            if dt_val < datetime.date(1900, 3, 1):
                epoch = datetime.date(1899, 12, 31)
            else:
                epoch = datetime.date(1899, 12, 30)
            delta = dt_val - epoch
            return float(delta.days)
    elif isinstance(dt_val, datetime.time):
        return (dt_val.hour * 3600 + dt_val.minute * 60 + dt_val.second + dt_val.microsecond / 1e6) / 86400.0
    return dt_val

def clean_and_normalize_coord(val, is_lat=True):
    if val is None:
        return None
    val = excel_datetime_to_float(val)
    try:
        val_str = str(val).strip().replace(',', '.')
        f_val = float(val_str)
        if f_val == 0:
            return 0.0
            
        sign = 1.0 if f_val >= 0 else -1.0
        abs_val = abs(f_val)
        
        if is_lat:
            # Lat in Vietnam [8.0, 25.0]
            if abs_val > 25.0:
                while abs_val > 25.0:
                    abs_val /= 10.0
            elif abs_val < 8.0 and abs_val > 0:
                while abs_val < 8.0:
                    abs_val *= 10.0
        else:
            # Lng in Vietnam [102.0, 110.0]
            if abs_val > 110.0:
                while abs_val > 110.0:
                    abs_val /= 10.0
            elif abs_val < 102.0 and abs_val > 0:
                while abs_val < 102.0:
                    abs_val *= 10.0
                    
        return sign * abs_val
    except (ValueError, TypeError):
        return None

def is_mock_coordinate(lat, lng):
    if lat is None or lng is None:
        return True
    try:
        f_lat = float(lat)
        f_lng = float(lng)
        if f_lat == 0 or f_lng == 0:
            return True
        # Check Vietnam bounding box
        # Lat roughly [8.0, 25.0]
        # Lng roughly [102.0, 110.0]
        if not (8.0 <= f_lat <= 25.0) or not (102.0 <= f_lng <= 110.0):
            return True
    except (ValueError, TypeError):
        return True
        
    lat_str = str(lat).strip()
    lng_str = str(lng).strip()
    if re.search(r'\.\d*85$', lat_str) and re.search(r'\.\d*42$', lng_str):
        return True
    return False

def load_coords_cache(store_data_js_path):
    cache = {}
    if os.path.exists(store_data_js_path):
        try:
            with open(store_data_js_path, "r", encoding="utf-8-sig") as f:
                content = f.read().replace("const STORE_LIST_DATA = ", "").strip(" \r\n;")
                data = json.loads(content)
                for s in data:
                    sid = str(s.get("id")).strip()
                    coords = s.get("coords")
                    if coords and coords.get("lat") and coords.get("lng"):
                        lat = float(coords["lat"])
                        lng = float(coords["lng"])
                        if not is_mock_coordinate(lat, lng):
                            cache[sid] = (lat, lng)
            print(f"Loaded {len(cache)} coordinates from store_data.js cache.")
        except Exception as e:
            print(f"Warning: Could not load coordinates cache: {e}")
    return cache

def main():
    print("Local DS_Cua_Hang.xlsm is now the sole source of truth (Google Sheet download disabled).")
            
    print("Loading Con Cung website store JSON...")
    if not os.path.exists(concung_json_path):
        print(f"Error: JSON file not found at {concung_json_path}")
        sys.exit(1)
        
    with open(concung_json_path, "r", encoding="utf-8-sig") as f:
        cc_stores = json.load(f)
        
    cc_map = {str(s["id_store"]).strip(): s for s in cc_stores}
    
    # Manual store details overrides for known website geocoding or name mismatches
    STORE_DETAILS_OVERRIDES = {
        "1132": {
            "name": "VPH - 118 Ngô Quyền",
            "address": "118 Ngô Quyền, phường Ngô Quyền, Thành phố Vĩnh Yên, Vĩnh Phúc",
            "province": "Tỉnh Vĩnh Phúc",
            "lat": 21.3077283,
            "lng": 105.5935362
        },
        "1212": {
            "name": "BNI - Thửa 163 Thôn Đại Thượng",
            "address": "Thôn Đại Thượng, xã Đại Đồng, huyện Tiên Du, Bắc Ninh",
            "province": "Tỉnh Bắc Ninh",
            "lat": 21.089416503906,
            "lng": 105.98628997803
        },
        "1191": {
            "name": "BVT - 166 Nguyễn Văn Cừ",
            "address": "166 Nguyễn Văn Cừ, phường Long Toàn, Thành phố Bà Rịa, Bà Rịa - Vũng Tàu",
            "province": "Tỉnh Bà Rịa - Vũng Tàu",
            "lat": 10.4927628,
            "lng": 107.1954681
        },
        "1210": {
            "name": "HYE- 190 đường ĐT179",
            "address": "190 đường ĐT179, Thị trấn Văn Giang, huyện Văn Giang, Hưng Yên",
            "province": "Tỉnh Hưng Yên",
            "lat": 20.93282,
            "lng": 105.93297
        },
        "1207": {
            "name": "BDI - 508 Quang Trung",
            "address": "508 Quang Trung, Thị trấn Phù Mỹ, huyện Phù Mỹ, Bình Định",
            "province": "Tỉnh Bình Định",
            "lat": 14.174611091614,
            "lng": 109.05133056641
        },
        "1208": {
            "name": "QNG - Quốc lộ 1A Tịnh Phong",
            "address": "Quốc lộ 1A, xã Tịnh Phong, huyện Sơn Tịnh, Quảng Ngãi",
            "province": "Tỉnh Quảng Ngãi",
            "lat": 15.18957901001,
            "lng": 108.79389953613
        },
        "1209": {
            "name": "NAN - 12-14 Trần Quốc Hoàn",
            "address": "12-14 Trần Quốc Hoàn, Khối Tân Thành, phường Hoà Hiếu, Thị xã Thái Hoà, Nghệ An",
            "province": "Tỉnh Nghệ An",
            "lat": 19.322343826294,
            "lng": 105.43447113037
        },
        "1213": {
            "name": "QNA - Đường ĐT 609C Đại Minh",
            "address": "Đường ĐT 609C, xã Đại Minh, huyện Đại Lộc, Quảng Nam",
            "province": "Tỉnh Quảng Nam",
            "lat": 15.849688529968,
            "lng": 108.06721496582
        },
        "1214": {
            "name": "KGI - 58-60 Đường 3/2",
            "address": "58 - 60 Đường 3/2, Khu phố Phước Trung 2, Thị trấn Gò Quao, huyện Gò Quao, Kiên Giang",
            "province": "Tỉnh Kiên Giang",
            "lat": 9.7291097640991,
            "lng": 105.27548980713
        },
        "1215": {
            "name": "DON - Đường Nguyễn Hoàng Sông Trầu",
            "address": "Đường Nguyễn Hoàng, Tổ 14, Ấp 2, xã Sông Trầu, huyện Trảng Bom, Đồng Nai",
            "province": "Tỉnh Đồng Nai",
            "lat": 10.976610183716,
            "lng": 107.02288818359
        },
        "1216": {
            "name": "NDI - Quốc lộ 37B Vụ Bản",
            "address": "Quốc lộ 37B, Tổ dân phố Văn Côi, Thị trấn Gôi, huyện Vụ Bản, Nam Định",
            "province": "Tỉnh Nam Định",
            "lat": 20.332136154175,
            "lng": 106.08071136475
        },
        "1217": {
            "name": "LDO - 262 Lê Lợi Đạ M'ri",
            "address": "262 Lê Lợi, Tổ dân phố 3, Thị trấn Đạ M'ri, huyện Đạ Huoai, Lâm Đồng",
            "province": "Tỉnh Lâm Đồng",
            "lat": 11.410870552063,
            "lng": 107.65926361084
        },
        "1219": {
            "name": "LDO - 30 Quốc Lộ 20 Di Linh",
            "address": "30 Quốc Lộ 20, thôn Đồng Lạc 1, xã Đinh Lạc, huyện Di Linh, Lâm Đồng",
            "province": "Tỉnh Lâm Đồng",
            "lat": 11.5180736,
            "lng": 108.1097512
        },
        "1220": {
            "name": "TNI - Đường HL8 Phước Bình",
            "address": "Đường HL8, ấp Bình Hòa, xã Phước Bình, Thị xã Trảng Bàng, Tây Ninh",
            "province": "Tỉnh Tây Ninh",
            "lat": 11.024748802185,
            "lng": 106.22874450684
        },
        "1221": {
            "name": "HDU - 23 Nguyễn Lương Bằng",
            "address": "23 Nguyễn Lương Bằng, Khu 3, Thị trấn Ninh Giang, huyện Ninh Giang, Hải Dương",
            "province": "Tỉnh Hải Dương",
            "lat": 20.7337589263916,
            "lng": 106.393257141113
        },
        "1298": {
            "name": "BTH - Số 82 đường Trung Tâm",
            "address": "Số 82, Đường Trung Tâm, thôn Mê Pu 2, xã Nam Thành, Lâm Đồng",
            "province": "Tỉnh Bình Thuận",
            "lat": 11.238067626953,
            "lng": 107.61563873291
        },
        "1319": {
            "lat": 20.7452550,
            "lng": 106.0610918
        },
        "1411": {
            "name": "HAG - 329A Nguyễn Trãi",
            "address": "329A Nguyễn Trãi, phường Nguyễn Trãi, Thành phố Hà Giang, Tỉnh Hà Giang",
            "province": "Tỉnh Hà Giang",
            "lat": 22.821105957031,
            "lng": 104.98170471191
        },
        "1415": {
            "name": "DLK - 91 Nguyễn Tất Thành",
            "address": "91 Nguyễn Tất Thành, Tổ 3A, xã Ea Kar, Đăk Lắk",
            "province": "Tỉnh Đắk Lắk",
            "lat": 12.8031278,
            "lng": 108.4730509
        },
        "1422": {
            "name": "DLK - Thôn 5 - Ea Kiết",
            "address": "Thôn 5, xã Ea Kiết, Đắk Lắk",
            "province": "Tỉnh Đắk Lắk",
            "lat": 12.8959234,
            "lng": 108.0696537
        },
        "1423": {
            "name": "VLO - 3 tháng 2 Tân Quới",
            "address": "Đường 3 Tháng 2, xã Tân Quới, Vĩnh Long",
            "province": "Tỉnh Vĩnh Long",
            "lat": 10.1293974,
            "lng": 105.7782468
        },
        "1425": {
            "name": "BTH - 514 Thủ Khoa Huân",
            "address": "514 Thủ Khoa Huân, phường Phú Thủy, Phan Thiết, Bình Thuận",
            "province": "Tỉnh Bình Thuận",
            "lat": 10.934958457947,
            "lng": 108.12947082519
        },
        "1426": {
            "lat": 21.1216607,
            "lng": 105.9655750
        },
        "1429": {
            "name": "THO - Minh Thịnh - Vạn Lộc",
            "address": "thôn Minh Thịnh, xã Vạn Lộc, Thanh Hoá",
            "province": "Tỉnh Thanh Hóa",
            "lat": 19.9330073,
            "lng": 105.9546771
        },
        "1430": {
            "name": "HCM - 111 Tân Chánh Hiệp 33",
            "address": "111 Tân Chánh Hiệp 33, phường Trung Mỹ Tây, Thành phố Hồ Chí Minh",
            "province": "Thành Phố Hồ Chí Minh",
            "lat": 10.8612652,
            "lng": 106.6642083
        },
        "1437": {
            "name": "GLA - 469 Hùng Vương",
            "address": "469 Hùng Vương, xã Phú Thiện, Gia Lai",
            "province": "Tỉnh Gia Lai",
            "lat": 13.5066242,
            "lng": 108.3271823
        },
        "1441": {
            "name": "HNI - 251 TL417",
            "address": "251 TL417, Liên Minh, Hà Nội 100000, Việt Nam",
            "province": "Thành Phố Hà Nội",
            "lat": 21.1210897,
            "lng": 105.6365972
        }
    }
    for osid, odata in STORE_DETAILS_OVERRIDES.items():
        if osid in cc_map:
            if "lat" in odata: cc_map[osid]["latitude"] = str(odata["lat"])
            if "lng" in odata: cc_map[osid]["longitude"] = str(odata["lng"])
            print(f"Applied coordinate override for store ID {osid}: {odata.get('lat')}, {odata.get('lng')}")
    print(f"Loaded {len(cc_stores)} stores from Con Cung JSON.")
    
    # Load cache coordinates
    cache = load_coords_cache(store_data_js_path)
    
    # Load local xlsm data as the source of truth
    print(f"Reading data from local xlsm {LOCAL_EXCEL_PATH}...")
    if not os.path.exists(LOCAL_EXCEL_PATH):
        print(f"Error: Local workbook {LOCAL_EXCEL_PATH} not found. Exiting.")
        sys.exit(1)
    gs_wb = openpyxl.load_workbook(LOCAL_EXCEL_PATH, data_only=True)
    gs_sheet = gs_wb["DS Cua Hang"]
    
    # Unmerge to read correctly
    if gs_sheet.merged_cells.ranges:
        for merged_range in list(gs_sheet.merged_cells.ranges):
            gs_sheet.unmerge_cells(str(merged_range))
            
    header_row = [cell.value for cell in gs_sheet[1]]
    col_map = {name: idx for idx, name in enumerate(header_row)}
    print("GSheet columns:", col_map)
    
    # Track template styling from the first data row
    template_row_style = []
    for col_idx in range(1, gs_sheet.max_column + 1):
        cell = gs_sheet.cell(row=2, column=col_idx)
        template_row_style.append({
            'font': copy(cell.font) if cell.font else None,
            'border': copy(cell.border) if cell.border else None,
            'fill': copy(cell.fill) if cell.fill else None,
            'alignment': copy(cell.alignment) if cell.alignment else None,
            'number_format': cell.number_format
        })
    
    processed_rows = []
    sheet_ids = set()
    error_count = 0
    
    for r_idx in range(2, gs_sheet.max_row + 1):
        id_val = gs_sheet.cell(row=r_idx, column=col_map['Ma Cua Hang (ID)'] + 1).value
        if id_val is None:
            continue
        sid = str(int(float(id_val))) if isinstance(id_val, (int, float)) else str(id_val).strip()
        sheet_ids.add(sid)
        
        # Handle manual mappings for GSheet IDs corresponding to different Web Store IDs
        lookup_sid = sid
        if sid in ["869", "942"]:
            lookup_sid = "441"
            sheet_ids.add("441")
        
        name_val = gs_sheet.cell(row=r_idx, column=col_map['Ten Cua Hang (Name)'] + 1).value
        addr_val = gs_sheet.cell(row=r_idx, column=col_map['Dia Chi (Address)'] + 1).value
        prov_val = gs_sheet.cell(row=r_idx, column=col_map['Tinh (Province)'] + 1).value
        
        raw_lat = gs_sheet.cell(row=r_idx, column=col_map['Vi Do (Lat)'] + 1).value
        raw_lng = gs_sheet.cell(row=r_idx, column=col_map['Kinh Do (Lng)'] + 1).value
        
        lat_val = clean_and_normalize_coord(raw_lat, is_lat=True)
        lng_val = clean_and_normalize_coord(raw_lng, is_lat=False)
        sheet_mock = is_mock_coordinate(lat_val, lng_val)
        
        is_error = False
        final_lat = lat_val
        final_lng = lng_val
        web_addr = None
        web_lat = None
        web_lng = None
        status_val = "Da doi chieu"
        
        if lookup_sid in cc_map:
            cc_store = cc_map[lookup_sid]
            cc_lat = float(cc_store["latitude"])
            cc_lng = float(cc_store["longitude"])
            cc_addr = cc_store["address"]
            
            web_addr = cc_addr
            web_lat = cc_lat
            web_lng = cc_lng
            
            if is_mock_coordinate(cc_lat, cc_lng):
                # Web coordinates are mock! Use sheet or cache if valid
                if not sheet_mock:
                    final_lat = lat_val
                    final_lng = lng_val
                    is_error = False
                elif lookup_sid in cache:
                    final_lat = cache[lookup_sid][0]
                    final_lng = cache[lookup_sid][1]
                    is_error = False
                elif sid in cache:
                    final_lat = cache[sid][0]
                    final_lng = cache[sid][1]
                    is_error = False
                else:
                    final_lat = cc_lat
                    final_lng = cc_lng
                    is_error = True
            else:
                # Web has valid coordinates. Use them, unless sheet coordinates are already valid
                if not sheet_mock:
                    final_lat = lat_val
                    final_lng = lng_val
                else:
                    final_lat = cc_lat
                    final_lng = cc_lng
                name_val = cc_store.get('name_support') or cc_store.get('name') or name_val
                addr_val = cc_addr
                prov_val = clean_province_name(cc_store.get('province_name') or prov_val, name_val)
                try:
                    if not sheet_mock:
                        is_error = False
                    elif sheet_mock or abs(lat_val - cc_lat) > 1e-5 or abs(lng_val - cc_lng) > 1e-5:
                        if sid in ["869", "942"]:
                            is_error = False
                        else:
                            is_error = True
                    else:
                        is_error = False
                except:
                    is_error = True
        else:
            status_val = "KhÃ´ng tÃ¬m tháº¥y trÃªn web"
            prov_val = clean_province_name(prov_val, name_val)
            if not sheet_mock:
                final_lat = lat_val
                final_lng = lng_val
                is_error = False
            elif sid in cache:
                final_lat = cache[sid][0]
                final_lng = cache[sid][1]
                is_error = False
            else:
                final_lat = lat_val
                final_lng = lng_val
                is_error = True
                
        # Apply manual details overrides
        override_id = lookup_sid if lookup_sid in STORE_DETAILS_OVERRIDES else (sid if sid in STORE_DETAILS_OVERRIDES else None)
        if override_id:
            overrides = STORE_DETAILS_OVERRIDES[override_id]
            if "name" in overrides: name_val = overrides["name"]
            if "address" in overrides: addr_val = overrides["address"]
            if "province" in overrides: prov_val = overrides["province"]
            if "lat" in overrides: final_lat = overrides["lat"]
            if "lng" in overrides: final_lng = overrides["lng"]
            is_error = False
            status_val = "Da doi chieu"
            
        if is_error:
            error_count += 1
            
        processed_rows.append({
            'id': int(sid),
            'name': name_val,
            'address': addr_val,
            'province': prov_val,
            'lat': final_lat,
            'lng': final_lng,
            'web_addr': web_addr,
            'web_lat': web_lat,
            'web_lng': web_lng,
            'status': status_val,
            'is_error': is_error
        })
        
    print(f"Reconciled sheet data. Found {error_count} stores with inaccurate coordinates or not found.")
    
    # Check for missing stores in Con Cung JSON (not in GSheet) and append
    missing_ids = [sid for sid in cc_map.keys() if sid not in sheet_ids]
    print(f"Found {len(missing_ids)} missing stores to append.")
    for sid in missing_ids:
        cc_store = cc_map[sid]
        cc_lat = float(cc_store["latitude"])
        cc_lng = float(cc_store["longitude"])
        cc_name = cc_store.get("name_support") or cc_store["name"]
        cc_addr = cc_store["address"]
        cc_prov = clean_province_name(cc_store["province_name"], cc_name)
        
        # Check if coordinates are mock
        is_mock = is_mock_coordinate(cc_lat, cc_lng)
        final_lat = cc_lat
        final_lng = cc_lng
        if is_mock and sid in cache:
            final_lat = cache[sid][0]
            final_lng = cache[sid][1]
            is_mock = False
            
        # Apply manual details overrides
        if sid in STORE_DETAILS_OVERRIDES:
            overrides = STORE_DETAILS_OVERRIDES[sid]
            if "name" in overrides: cc_name = overrides["name"]
            if "address" in overrides: cc_addr = overrides["address"]
            if "province" in overrides: cc_prov = overrides["province"]
            if "lat" in overrides: final_lat = overrides["lat"]
            if "lng" in overrides: final_lng = overrides["lng"]
            is_mock = False
            
        processed_rows.append({
            'id': int(sid),
            'name': cc_name,
            'address': cc_addr,
            'province': cc_prov,
            'lat': final_lat,
            'lng': final_lng,
            'web_addr': cc_addr,
            'web_lat': cc_lat,
            'web_lng': cc_lng,
            'status': "Da doi chieu",
            'is_error': is_mock  # New stores with mock coords are marked as errors
        })
        
    # Sort: Errors first, then ID ascending
    processed_rows.sort(key=lambda x: (0 if x['is_error'] else 1, x['id']))
    
    # Close gsheet workbook
    gs_wb.close()
    
    # 1. Save sorted clean workbook to DS_Cua_Hang_Updated.xlsx (10 columns, no Region)
    print(f"Saving updated clean workbook to {updated_xlsx_path}...")
    up_wb = openpyxl.Workbook()
    up_sheet = up_wb.active
    up_sheet.title = "DS Cua Hang"
    
    # Write headers
    up_headers = ['Ma Cua Hang (ID)', 'Ten Cua Hang (Name)', 'Dia Chi (Address)', 'Tinh (Province)', 
                  'Vi Do (Lat)', 'Kinh Do (Lng)', 'Äá»‹a chá»‰ (Website)', 'Lat (Website)', 'Lng (Website)', 'Tráº¡ng thÃ¡i Ä‘á»‘i chiáº¿u']
    up_sheet.append(up_headers)
    
    for r_idx, row_data in enumerate(processed_rows, start=2):
        up_sheet.cell(row=r_idx, column=1, value=row_data['id'])
        up_sheet.cell(row=r_idx, column=2, value=row_data['name'])
        up_sheet.cell(row=r_idx, column=3, value=row_data['address'])
        up_sheet.cell(row=r_idx, column=4, value=row_data['province'])
        up_sheet.cell(row=r_idx, column=5, value=row_data['lat'])
        up_sheet.cell(row=r_idx, column=6, value=row_data['lng'])
        up_sheet.cell(row=r_idx, column=7, value=row_data['web_addr'])
        up_sheet.cell(row=r_idx, column=8, value=row_data['web_lat'])
        up_sheet.cell(row=r_idx, column=9, value=row_data['web_lng'])
        up_sheet.cell(row=r_idx, column=10, value=row_data['status'])
        
        # Apply style from template
        for col_idx in range(1, 11):
            cell = up_sheet.cell(row=r_idx, column=col_idx)
            # Find template style matching this or use first column style
            tpl_style = template_row_style[col_idx-1] if col_idx-1 < len(template_row_style) else template_row_style[0]
            if tpl_style['font']: cell.font = copy(tpl_style['font'])
            if tpl_style['border']: cell.border = copy(tpl_style['border'])
            if tpl_style['fill']: cell.fill = copy(tpl_style['fill'])
            if tpl_style['alignment']: cell.alignment = copy(tpl_style['alignment'])
            cell.number_format = tpl_style['number_format']
            
    up_wb.save(updated_xlsx_path)
    up_wb.close()
    print("Saved clean workbook successfully.")
    
    # 2. Save sorted workbook to DS_Cua_Hang.xlsm (11 columns, including Region and preserving VBA)
    print(f"Loading local xlsm workbook from {LOCAL_EXCEL_PATH}...")
    if not os.path.exists(LOCAL_EXCEL_PATH):
        print(f"Error: Local workbook {LOCAL_EXCEL_PATH} not found.")
        sys.exit(1)
        
    xl_wb = openpyxl.load_workbook(LOCAL_EXCEL_PATH, keep_vba=True)
    xl_sheet = xl_wb["DS Cua Hang"]
    
    # Unmerge to write safely
    if xl_sheet.merged_cells.ranges:
        for merged_range in list(xl_sheet.merged_cells.ranges):
            xl_sheet.unmerge_cells(str(merged_range))
            
    # Clear original data rows 2 to max_row
    max_row = xl_sheet.max_row
    for r in range(2, max_row + 1):
        for c in range(1, xl_sheet.max_column + 1):
            xl_sheet.cell(row=r, column=c).value = None
            
    # Write updated rows
    for r_idx, row_data in enumerate(processed_rows, start=2):
        is_north = is_northern_province(row_data['province'])
        region_val = "North" if is_north else None
        
        xl_sheet.cell(row=r_idx, column=1, value=row_data['id'])
        xl_sheet.cell(row=r_idx, column=2, value=row_data['name'])
        xl_sheet.cell(row=r_idx, column=3, value=row_data['address'])
        xl_sheet.cell(row=r_idx, column=4, value=row_data['province'])
        xl_sheet.cell(row=r_idx, column=5, value=region_val)
        xl_sheet.cell(row=r_idx, column=6, value=row_data['lat'])
        xl_sheet.cell(row=r_idx, column=7, value=row_data['lng'])
        xl_sheet.cell(row=r_idx, column=8, value=row_data['web_addr'])
        xl_sheet.cell(row=r_idx, column=9, value=row_data['web_lat'])
        xl_sheet.cell(row=r_idx, column=10, value=row_data['web_lng'])
        xl_sheet.cell(row=r_idx, column=11, value=row_data['status'])
        
        # Apply styling
        for col_idx in range(1, 12):
            cell = xl_sheet.cell(row=r_idx, column=col_idx)
            tpl_idx = col_idx - 1 if col_idx < 5 else (col_idx - 2 if col_idx > 5 else 3)
            tpl_style = template_row_style[tpl_idx] if tpl_idx < len(template_row_style) else template_row_style[0]
            if tpl_style['font']: cell.font = copy(tpl_style['font'])
            if tpl_style['border']: cell.border = copy(tpl_style['border'])
            if tpl_style['fill']: cell.fill = copy(tpl_style['fill'])
            if tpl_style['alignment']: cell.alignment = copy(tpl_style['alignment'])
            cell.number_format = tpl_style['number_format']
            
    xl_wb.save(LOCAL_EXCEL_PATH)
    xl_wb.close()
    print("Saved local xlsm workbook successfully.")
    print("\n--- Python Store Update Complete! ---")

if __name__ == "__main__":
    main()

