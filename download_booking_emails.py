import imaplib
import email
from email.header import decode_header
import os
import re
import sys
import hashlib
import io

sys.stdout.reconfigure(encoding='utf-8')

# ========================================================
# CẤU HÌNH TÀI KHOẢN EMAIL VÀ THƯ MỤC LƯU TRỮ
# ========================================================
IMAP_SERVER = "imap.gmail.com"
EMAIL_ACCOUNT = "cuongnd@ghn.vn"

# BẠN CẦN ĐIỀN MẬT KHẨU ỨNG DỤNG VÀO FILE .env (EMAIL_APP_PASSWORD)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

APP_PASSWORD = os.environ.get("EMAIL_APP_PASSWORD", "")

# Thư mục lưu file Excel Booking tải về
OUTPUT_DIR = r"g:\My Drive\Training AI\Supra Phú Thọ\Data_Booking"

# Tiêu đề email cần tìm kiếm chính xác
TARGET_SUBJECT = "DCPT"
# ========================================================

def decode_mime_header(header_value):
    if not header_value:
        return ""
    decoded_parts = decode_header(header_value)
    result = []
    for text, encoding in decoded_parts:
        if isinstance(text, bytes):
            try:
                result.append(text.decode(encoding if encoding else "utf-8", errors="ignore"))
            except Exception:
                result.append(text.decode("latin1", errors="ignore"))
        else:
            result.append(str(text))
    return "".join(result)

def get_file_hash(path):
    if not os.path.exists(path):
        return ""
    try:
        with open(path, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()
    except Exception:
        return ""

def download_attachments():
    if not APP_PASSWORD:
        print("[LỖI] Vui lòng cấu hình EMAIL_APP_PASSWORD trong file .env trước khi chạy!")
        return

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"Đã tạo thư mục lưu trữ: {OUTPUT_DIR}")

    try:
        print(f"Đang kết nối tới máy chủ IMAP: {IMAP_SERVER}...")
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_ACCOUNT, APP_PASSWORD)
        print("Đăng nhập tài khoản email thành công!")
        
        # Tự động tìm thư mục "All Mail" (Tất cả thư) của Gmail để tránh bỏ sót các email đã lưu trữ/lọc ngoài Inbox
        all_mail_folder = "inbox"
        try:
            status, folder_list = mail.list()
            if status == "OK":
                for folder_info in folder_list:
                    folder_str = folder_info.decode('utf-8', errors='ignore')
                    if '\\All' in folder_str:
                        parts = folder_str.split(' "/" ')
                        if len(parts) == 2:
                            all_mail_folder = parts[1].strip().strip('"')
                            break
        except Exception as folder_err:
            print(f"Không thể lấy danh sách thư mục, sử dụng Inbox mặc định: {folder_err}")
            
        print(f"Đang mở thư mục: {all_mail_folder}...")
        mail.select(f'"{all_mail_folder}"')
        
        # Tìm kiếm email bằng tiêu đề chính xác (sử dụng IMAP literal để hỗ trợ ký tự UTF-8 tiếng Việt có dấu)
        print("Đang tìm kiếm email...")
        email_ids = []
        
        # Tìm kiếm DCPT
        mail.literal = "DCPT".encode('utf-8')
        status1, messages1 = mail.search('UTF-8', 'SUBJECT')
        if status1 == "OK" and messages1[0]:
            email_ids.extend(messages1[0].split())
            
        # Loại bỏ các ID trùng lặp và sắp xếp theo ID (từ cũ đến mới)
        email_ids = sorted(list(set(email_ids)), key=int)
        
        if not email_ids:
            print("Không tìm thấy email nào phù hợp với từ khóa.")
            return
        print(f"Tìm thấy tổng cộng {len(email_ids)} email phù hợp.")
        
        download_count = 0
        processed_booking_files = set()
        processed_planning_dates = set()
        
        for e_id in reversed(email_ids[-300:]):
            _, msg_data = mail.fetch(e_id, "(RFC822)")
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    
                    subject = decode_mime_header(msg["Subject"])
                    sender = decode_mime_header(msg["From"])
                    
                    # Kiểm tra chính xác tiêu đề chứa từ khóa DCPT (không phân biệt chữ hoa/thường)
                    is_valid_subject = "DCPT".upper() in subject.upper()
                    if not is_valid_subject:
                        continue
                        
                    print(f"\nĐang kiểm tra Email: \"{subject}\" từ {sender}")
                    
                    # Lấy ngày booking từ tiêu đề để chuẩn hóa tên file Planning
                    booking_date = None
                    date_match = re.search(r'(\d{1,2})\s*[\./-]\s*(\d{1,2})\s*[\./-]\s*(\d{4})', subject)
                    if not date_match:
                        # Thử định dạng YYYY.MM.DD
                        date_match_rev = re.search(r'(\d{4})\s*[\./-]\s*(\d{1,2})\s*[\./-]\s*(\d{1,2})', subject)
                        if date_match_rev:
                            y, m, d = date_match_rev.groups()
                            booking_date = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                    else:
                        d, m, y = date_match.groups()
                        booking_date = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                    
                    # Duyệt các phần đính kèm trong email
                    if msg.is_multipart():
                        for part in msg.walk():
                            content_disposition = str(part.get("Content-Disposition"))
                            filename = part.get_filename()
                            
                            # Nếu có file đính kèm
                            if filename and "attachment" in content_disposition:
                                filename = decode_mime_header(filename)
                                # Chuẩn hóa tên file (loại bỏ ký tự lạ nếu có)
                                filename = re.sub(r'[\/*?:"<>|]', "", filename)
                                print(f"    - Đính kèm: {filename}")
                                
                                # Chỉ tải các file Excel (.xlsx, .xls, .xlsb)
                                if filename.lower().endswith(('.xlsx', '.xls', '.xlsb')):
                                    is_planning = filename.upper().startswith("PLANNING")
                                    is_booking = not is_planning
                                    
                                    # Try to extract date from filename first
                                    file_date = None
                                    # Pattern: DD.MM.YYYY or DD-MM-YYYY (e.g. 26.6.2026)
                                    f_match = re.search(r'(\d{1,2})[\.\-](\d{1,2})[\.\-](\d{4})', filename)
                                    if f_match:
                                        fd, fm, fy = f_match.groups()
                                        file_date = f"{fy}-{fm.zfill(2)}-{fd.zfill(2)}"
                                    else:
                                        # Pattern: DD.MM or DD-MM (e.g. 26.6)
                                        f_match2 = re.search(r'(\d{1,2})[\.\-](\d{1,2})', filename)
                                        if f_match2:
                                            fd, fm = f_match2.groups()
                                            if 1 <= int(fd) <= 31 and 1 <= int(fm) <= 12:
                                                file_date = f"2026-{fm.zfill(2)}-{fd.zfill(2)}"
                                    
                                    active_booking_date = file_date if file_date else booking_date
                                    
                                    if True:
                                        if is_planning and active_booking_date:
                                            target_filename = f"Planning_{active_booking_date}.xlsx"
                                            dest_path = os.path.join(OUTPUT_DIR, target_filename)
                                            
                                            is_ghn = "ghn.vn" in sender.lower()
                                            sender_type = "GHN" if is_ghn else "SUPRA"
                                            
                                            # Lấy dữ liệu file đính kèm
                                            file_data = part.get_payload(decode=True)
                                            new_hash = hashlib.md5(file_data).hexdigest()
                                            
                                            # Đếm số biển số xe trong file Planning đính kèm
                                            new_plates = 0
                                            try:
                                                import openpyxl
                                                wb_check = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
                                                sh = wb_check['Form'] if 'Form' in wb_check.sheetnames else wb_check.active
                                                for mr in list(sh.merged_cells.ranges):
                                                    sh.unmerge_cells(str(mr))
                                                for row_idx in range(3, sh.max_row + 1):
                                                    plate_val = sh.cell(row=row_idx, column=12).value
                                                    if plate_val and str(plate_val).strip():
                                                        new_plates += 1
                                                wb_check.close()
                                            except Exception as e:
                                                print(f" -> Lỗi đếm biển số xe: {e}")
                                            
                                            # Lấy thông tin file trên đĩa
                                            disk_plates = 0
                                            disk_hash = ""
                                            if os.path.exists(dest_path):
                                                disk_hash = get_file_hash(dest_path)
                                                try:
                                                    wb_disk = openpyxl.load_workbook(dest_path, data_only=True)
                                                    sh_disk = wb_disk['Form'] if 'Form' in wb_disk.sheetnames else wb_disk.active
                                                    for mr in list(sh_disk.merged_cells.ranges):
                                                        sh_disk.unmerge_cells(str(mr))
                                                    for row_idx in range(3, sh_disk.max_row + 1):
                                                        plate_val = sh_disk.cell(row=row_idx, column=12).value
                                                        if plate_val and str(plate_val).strip():
                                                            disk_plates += 1
                                                    wb_disk.close()
                                                except Exception:
                                                    pass
                                            
                                            # Quyết định tải file
                                            should_download = False
                                            reason = ""
                                            
                                            if active_booking_date not in processed_planning_dates:
                                                processed_planning_dates.add(active_booking_date)
                                                
                                                if not os.path.exists(dest_path):
                                                    should_download = True
                                                    reason = f"File chưa tồn tại cục bộ"
                                                elif new_hash != disk_hash:
                                                    # Ưu tiên bản mới nhất nếu nội dung thay đổi, 
                                                    # trừ khi bản mới 0 có biển số nhưng bản cũ có (tránh ghi đè nhầm do reply mail đính kèm file gốc)
                                                    if new_plates > 0 or disk_plates == 0:
                                                        should_download = True
                                                        reason = f"Bản mới nhất có cập nhật nội dung (khác hash, {new_plates} biển số)"
                                                    else:
                                                        reason = "Bản mới nhất không có biển số, giữ lại bản cũ có biển số để tránh mất dữ liệu"
                                                else:
                                                    reason = "Bản trên đĩa đã trùng khớp hoàn toàn"
                                            else:
                                                if new_plates > disk_plates:
                                                    should_download = True
                                                    reason = f"Bản cũ hơn từ email có nhiều biển số hơn bản hiện tại ({new_plates} > {disk_plates})"
                                            
                                            if should_download:
                                                print(f" -> Tải file Planning cho {active_booking_date} ({sender_type}, {new_plates} biển số): {filename}")
                                                print(f"    Chi tiết: {reason}")
                                                with open(dest_path, "wb") as f:
                                                    f.write(file_data)
                                                download_count += 1
                                            else:
                                                # Tránh in quá nhiều nếu trùng khớp
                                                if new_plates < disk_plates or new_hash != disk_hash:
                                                    print(f" -> Bỏ qua file Planning cho {active_booking_date} ({new_plates} biển số): {filename}. Chi tiết: {reason}")
                                        
                                        elif is_booking:
                                            if active_booking_date:
                                                y, m, d = active_booking_date.split('-')
                                                ext = os.path.splitext(filename)[1]
                                                target_filename = f"Booking Supra {d}-{m}-{y}{ext}"
                                            else:
                                                target_filename = filename
                                                
                                            if target_filename not in processed_booking_files:
                                                pass # just a placeholder for logic indentation
                                            
                                            dest_path = os.path.join(OUTPUT_DIR, target_filename)
                                            file_data = part.get_payload(decode=True)
                                            new_hash = hashlib.md5(file_data).hexdigest()
                                            
                                            if target_filename not in processed_booking_files:
                                                processed_booking_files.add(target_filename)
                                                disk_hash = get_file_hash(dest_path) if os.path.exists(dest_path) else ""
                                                
                                                if not os.path.exists(dest_path):
                                                    print(f" -> Phát hiện file Booking mới: {filename} (Lưu thành {target_filename}). Đang tải...")
                                                    with open(dest_path, "wb") as f:
                                                        f.write(file_data)
                                                    download_count += 1
                                                elif new_hash != disk_hash:
                                                    print(f" -> Phát hiện file Booking có nội dung cập nhật từ email: {filename} (Ghi đè {target_filename}). Đang tải...")
                                                    with open(dest_path, "wb") as f:
                                                        f.write(file_data)
                                                    download_count += 1
                                                else:
                                                    # Nội dung hoàn toàn trùng khớp
                                                    pass
                                            else:
                                                # Đã xử lý bản mới hơn của file Booking này trong cùng lượt chạy
                                                pass
                                        
        print(f"\n[HOÀN THÀNH] Đã tải về thêm {download_count} file Excel mới vào thư mục Data_Booking!")
        mail.close()
        mail.logout()
        
    except Exception as e:
        print(f"[LỖI] Đã xảy ra lỗi trong quá trình xử lý email: {e}")

if __name__ == "__main__":
    download_attachments()
