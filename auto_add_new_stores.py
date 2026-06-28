import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

def get_col_index(headers, col_name):
    for i, h in enumerate(headers):
        if h and isinstance(h, str) and col_name.lower() in h.lower():
            return i
    return -1

def main():
    folder_path = os.path.dirname(os.path.abspath(__file__))
    
    # DEBUG START
    try:
        debug_lines = []
        html_path = os.path.join(folder_path, "demo.html")
        if os.path.exists(html_path):
            html_content = open(html_path, encoding="utf-8").read()
            debug_lines.append("=== Searching for updateTripTables or table rendering ===")
            for idx, line in enumerate(html_content.split("\n")):
                if "updateTripTable" in line or "renderTrip" in line or "Lộ trình" in line or "routeDetail" in line or "points.map" in line:
                    if len(line.strip()) < 200:
                        debug_lines.append(f"Line {idx+1}: {line.strip()}")
        else:
            debug_lines.append("demo.html not found.")
            
        with open(os.path.join(folder_path, "scratch_debug_log.txt"), "w", encoding="utf-8") as df:
            df.write("\n".join(debug_lines))
    except Exception as de:
        with open(os.path.join(folder_path, "scratch_debug_log.txt"), "w", encoding="utf-8") as df:
            df.write(f"SEARCH EXCEPTION: {de}")
    # DEBUG END
    data_folder = os.path.join(folder_path, "Data_Booking")
    store_file = os.path.join(folder_path, "DSCuaHangFinal.xlsx")

    if not os.path.exists(store_file):
        print(f"Error: {store_file} not found.")
        return

    # 1. Load existing stores
    print("Loading existing stores from DSCuaHangFinal.xlsx...")
    try:
        wb_store = openpyxl.load_workbook(store_file)
        sheet_store = wb_store["Sheet1"]
    except Exception as e:
        print(f"Failed to load DS_Cua_Hang.xlsm: {e}")
        return

    existing_ids = set()
    for r in range(2, sheet_store.max_row + 1):
        val = sheet_store.cell(row=r, column=1).value
        if val is not None:
            try:
                existing_ids.add(int(float(val)))
            except ValueError:
                existing_ids.add(str(val).strip())

    new_stores = []
    new_ids_added_this_run = set()

    # 2. Iterate through booking files
    print("Scanning booking files for new stores...")
    # Pattern loại trừ file đã xếp tay (dùng regex để tránh vấn đề encoding Unicode)
    import re
    excluded_patterns = [
        r"DCMQ-CC.*ADHOC.*12\.06\.2026",
    ]
    def is_excluded(fname):
        return any(re.search(p, fname) for p in excluded_patterns)
    files = [f for f in os.listdir(data_folder) if f.endswith('.xlsx') and 'MQ-CC' in f and not f.startswith('~') and not is_excluded(f)]
    
    for file_name in files:
        file_path = os.path.join(data_folder, file_name)
        try:
            wb_booking = openpyxl.load_workbook(file_path, data_only=True)
            summary_sheet_name = None
            for name in wb_booking.sheetnames:
                if 'Summary' in name:
                    summary_sheet_name = name
                    break
            
            if not summary_sheet_name:
                wb_booking.close()
                continue
                
            sheet = wb_booking[summary_sheet_name]
            
            # Find header row
            header_row = -1
            headers = []
            for r in range(1, 10):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
                if any(v and isinstance(v, str) and "Ma CH" in v for v in row_vals):
                    header_row = r
                    headers = row_vals
                    break
                    
            if header_row == -1:
                wb_booking.close()
                continue
                
            col_id = get_col_index(headers, "Ma CH")
            col_name = get_col_index(headers, "TÃªn cá»­a hÃ ng")
            col_prov = get_col_index(headers, "Province")
            col_region = get_col_index(headers, "Region")
            
            if col_id == -1:
                wb_booking.close()
                continue
                
            for r in range(header_row + 1, sheet.max_row + 1):
                s_id_val = sheet.cell(row=r, column=col_id + 1).value
                if not s_id_val:
                    continue
                    
                try:
                    s_id = int(float(s_id_val))
                except ValueError:
                    s_id = str(s_id_val).strip()
                    
                if s_id not in existing_ids and s_id not in new_ids_added_this_run:
                    s_name = sheet.cell(row=r, column=col_name + 1).value if col_name != -1 else f"Store {s_id}"
                    s_prov = sheet.cell(row=r, column=col_prov + 1).value if col_prov != -1 else ""
                    s_region = sheet.cell(row=r, column=col_region + 1).value if col_region != -1 else ""
                    
                    # Extract district/address from store name to populate address
                    s_addr = s_name.split("-")[-1].strip() if s_name and "-" in s_name else s_name
                    if s_addr and s_prov:
                        s_addr = f"{s_addr}, {s_prov}"
                        
                    new_stores.append({
                        'id': s_id,
                        'name': s_name,
                        'addr': s_addr,
                        'prov': s_prov,
                        'region': s_region
                    })
                    new_ids_added_this_run.add(s_id)
            
            wb_booking.close()
            
        except Exception as e:
            print(f"Error processing file {file_name}: {e}")

    # 3. Append new stores
    if new_stores:
        print(f"Found {len(new_stores)} new stores. Appending to DSCuaHangFinal.xlsx...")
        start_row = sheet_store.max_row + 1
        for i, store in enumerate(new_stores):
            r = start_row + i
            sheet_store.cell(row=r, column=1, value=store['id'])
            sheet_store.cell(row=r, column=2, value=store['name'])
            sheet_store.cell(row=r, column=3, value=store['addr'])
            sheet_store.cell(row=r, column=4, value="")  # District (leave blank)
            sheet_store.cell(row=r, column=5, value=store['prov'])  # Province
            
        wb_store.save(store_file)
        print("Stores appended successfully.")
    else:
        print("No new stores found in booking files.")
        
    wb_store.close()

if __name__ == "__main__":
    main()

